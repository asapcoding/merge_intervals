'''
Created on 07.05.2020

@author: asapcoding
'''
'''libraries import / Bibliotheken einbinden'''
from openpyxl import load_workbook
from tkinter import filedialog
import numpy as np

'''functions / Funktionen'''
def loadfile():
    # opens dialog box to choose the excel-file
    # öffnet das Dialogfeld zur Auswahl der Excel-Datei
    file_path = filedialog.askopenfilename(title = "Exceldatei laden",filetypes = (("excel files","*.xlsx"),))
    
    # if: user selects a excel-file else: the user cancel the dialog box
    # if: Benutzer wählt eine Excel-Datei aus else: Der Benutzer bricht das Dialogfeld ab
    if file_path:
        # loads the excel-file into memory
        # lädt die Excel-Datei in den Speicher
        workbook = load_workbook(filename=file_path, data_only=True)
    
        # gets first sheet of excel-table
        # nimmt das erste Blatt der Excel-Tabelle
        sheet = workbook.active

        # gets the dimension of the sheet
        # ruft die Abmessung des Blattes ab
        sheetdim = sheet.dimensions
    
        # checks that the first cell is A1
        # prüft, ob die erste beschriebene Zelle A1 ist
        if sheetdim[0:2] == 'A1':
            
            # searches for B to ensure that the table has the given size
            # sucht nach B, um sicherzustellen, dass die Tabelle die vorgegebene Größe besitzt
            posB = sheetdim.find('B')
            
            # checks that the second column is B
            # prüft, ob die zweite Spalte B ist
            if  posB != -1:
                
                # splits the string into single character to get the number of letters
                # teilt die Zeichenfolge in ein einzelnes Zeichen auf, um die Anzahl der Buchstaben zu erhalten
                dimSplit = list(sheetdim)
                
                # gets the number of letters
                # erhält die Anzahl der Buchstaben
                dimEnd = len(dimSplit)
                
                # gets the number after B, join the character to one number
                # erhält die Nummer nach B, verbindet die Zeichen zu einer Zahl = Spaltenlänge
                lengthcol = int(''.join(dimSplit[posB+1:dimEnd]))

                # creates a matrix 2 x length of column
                # erstellt eine Matrix mit 2 x Spaltenlänge
                dataAB = np.zeros((lengthcol,2),dtype=int)
                
                # fill up the matrix with the value from excel-table
                # füllt die Matrix mit den Werten aus der Excel-Tabelle
                for x in range(1,lengthcol+1): 
                    
                    # gets values
                    # erhält Werte der Excel-Tabelle
                    dA = sheet.cell(row = x,column = 1)
                    dB = sheet.cell(row = x,column = 2)
                    
                    # write values into matrix
                    # schreibt Werte in die Matrix
                    dataAB[x-1,0] = dA.value
                    dataAB[x-1,1] = dB.value
                    
                # return the matrix
                # gibt die matrix zurück
                return dataAB
            else:
                # return 0 to flag an error
                # Null wird zurückgegeben, um einen Fehler zu kennzeichnen
                return 0
        else:
            # return 0 to flag an error
            # Null wird zurückgegeben, um einen Fehler zu kennzeichnen
            return 0
    else:
        # return 0 to flag an error
        # Null wird zurückgegeben, um einen Fehler zu kennzeichnen
        return 0
